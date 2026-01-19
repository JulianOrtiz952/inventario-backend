from rest_framework import viewsets, status, filters
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.response import Response
from rest_framework.exceptions import ValidationError
from decimal import Decimal, ROUND_HALF_UP
from django.db import transaction
from rest_framework.decorators import action
from django.db.models import Sum, Count, Q, F, Case, When, DecimalField, Value, IntegerField
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from .renderers import XLSXRenderer
import io
from datetime import datetime
from django.utils import timezone


from .models import (
    Insumo, Proveedor, Producto, Bodega, Impuesto, PrecioProducto,
    Tercero, DatosAdicionalesProducto, Talla,
    NotaEnsamble, ProductoInsumo, NotaEnsambleDetalle, NotaEnsambleInsumo,
    TrasladoProducto, NotaSalidaProducto, NotaSalidaAfectacionStock, InsumoMovimiento,
    ProductoTerminadoMovimiento
)
from .filters import InsumoFilter, ProductoFilter, NotaEnsambleFilter, NotaSalidaProductoFilter
from .serializers import (
    InsumoSerializer, ProveedorSerializer, ProductoSerializer, BodegaSerializer,
    ImpuestoSerializer, ProductoPrecioWriteSerializer,
    TerceroSerializer, DatosAdicionalesWriteSerializer,
    TallaSerializer, NotaEnsambleSerializer, ProductoInsumoSerializer,
    TrasladoProductoSerializer, NotaSalidaProductoSerializer, InsumoMovimientoSerializer, InsumoMovimientoInputSerializer,
    ProductoTerminadoMovimientoSerializer
)

def consumir_insumos_manuales_por_delta(nota, signo=Decimal("1")):
    """
    signo = +1 descuenta los insumos manuales asociados a la nota
    signo = -1 devuelve (reversa)
    """
    ins_to_update = []
    for ni in nota.insumos.all():
        cantidad = _d(ni.cantidad) * signo
        ins = ni.insumo
        if ins.bodega_id != nota.bodega_id:
            ins = Insumo.objects.filter(codigo=ni.insumo.codigo, bodega=nota.bodega).first()

        if not ins:
            raise ValidationError({"detail": f"Insumo {ni.insumo.codigo} no existe en la bodega de la nota."})

        if cantidad > 0 and _d(ins.cantidad) < cantidad:
             raise ValidationError({
                "stock_insuficiente": {
                    ins.codigo: {
                        "insumo": ins.nombre,
                        "disponible": str(_d(ins.cantidad)),
                        "requerido": str(_d(cantidad)),
                        "faltante": str(_d(cantidad) - _d(ins.cantidad)),
                    }
                }
            })

        ins.cantidad = _d(ins.cantidad) - cantidad
        ins_to_update.append(ins)

    if ins_to_update:
        Insumo.objects.bulk_update(ins_to_update, ["cantidad"])

def _decimal(v, field_name):
    try:
        return Decimal(str(v))
    except Exception:
        raise ValidationError({field_name: "Valor inválido"})

def registrar_movimiento_sin_afectar_stock(*, insumo, tercero, tipo, cantidad, costo_unitario, bodega=None, factura="", observacion="", nota_ensamble=None):
    """
    Registra historial SIN modificar stock (útil para CREACION cuando ya guardaste la cantidad en Insumo).
    """
    cantidad = _decimal(cantidad, "cantidad")
    if cantidad < 0:
        raise ValidationError({"cantidad": "No puede ser negativa"})

    costo_unitario = _decimal(costo_unitario, "costo_unitario")
    total = (cantidad * costo_unitario).quantize(Decimal("0.01"))

    mov = InsumoMovimiento.objects.create(
        insumo=insumo,
        tercero=tercero,
        bodega=bodega or insumo.bodega,
        tipo=tipo,
        cantidad=cantidad,
        unidad_medida=getattr(insumo, "unidad_medida", "") or "",
        costo_unitario=costo_unitario,
        total=total,
        saldo_resultante=insumo.cantidad,  # ya está guardado
        factura=factura or getattr(insumo, "factura", "") or "",
        observacion=observacion or "",
        nota_ensamble=nota_ensamble,
    )
    return mov

def aplicar_movimiento_insumo(*, insumo, tercero, tipo, cantidad, costo_unitario=None, bodega=None, factura="", observacion="", nota_ensamble=None):
    """
    Modifica stock + registra historial en una transacción.
    Para ENTRADA/SALIDA/AJUSTE/CONSUMO_ENSAMBLE.
    """
    cantidad = _decimal(cantidad, "cantidad")
    if cantidad <= 0:
        raise ValidationError({"cantidad": "Debe ser mayor a 0"})

    if costo_unitario is None or str(costo_unitario) == "":
        costo_unitario = insumo.costo_unitario or Decimal("0.00")
    else:
        costo_unitario = _decimal(costo_unitario, "costo_unitario")

    total = (cantidad * costo_unitario).quantize(Decimal("0.01"))

    with transaction.atomic():
        # 🔒 Bloquear el insumo para evitar modificaciones concurrentes
        # en lugar de insumo.refresh_from_db()
        insumo = Insumo.objects.select_for_update().get(pk=insumo.pk)

        if tipo in ("SALIDA", "CONSUMO_ENSAMBLE"):
            if insumo.cantidad < cantidad:
                raise ValidationError({"cantidad": "Stock global insuficiente"})
            
            # Validar stock de BODEGA específica (si se especifica bodega)
            if bodega:
                # 🔒 También necesitamos asegurar que el cálculo del stock de bodega sea consistente
                # Como InsumoMovimiento es append-only, no modificamos filas, pero leemos agregados.
                # Si alguien inserta un movimiento justo ahora, el agregado cambiará.
                # Idealmente deberíamos bloquear tablas o usar lógica de snapshots, pero
                # para este nivel, validar contra el insumo (que ya tiene lock) suele ser "suficiente"
                # si asumimos que la cantidad global es la fuente de verdad final, 
                # o si aceptamos cierto riesgo en la distribución por bodega.
                # 
                # SIN EMBARGO, para ser estrictos, deberíamos bloquear las filas de movimientos
                # o simplemente confiar en que el Insumo Lock serializa las escrituras sobre este insumo.
                # Al tener locked 'insumo', nadie más puede entrar a este bloque para este insumo,
                # así que nadie más insertará movimientos para ESTE insumo concurrentemente.
                # ¡Es seguro!
                
                qs = InsumoMovimiento.objects.filter(insumo=insumo, bodega=bodega)
                agg = qs.aggregate(
                    total_entradas=Sum(
                        Case(
                            When(tipo__in=["CREACION", "ENTRADA", "AJUSTE"], then=F("cantidad")),
                            default=0,
                            output_field=DecimalField()
                        )
                    ),
                    total_salidas=Sum(
                        Case(
                            When(tipo__in=["SALIDA", "CONSUMO_ENSAMBLE"], then=F("cantidad")),
                            default=0,
                            output_field=DecimalField()
                        )
                    )
                )
                entradas = agg["total_entradas"] or Decimal("0")
                salidas = agg["total_salidas"] or Decimal("0")
                stock_bodega = entradas - salidas
                
                if stock_bodega < cantidad:
                    raise ValidationError({"cantidad": f"Stock insuficiente en bodega {bodega.nombre}. Disponible: {stock_bodega}"})

            insumo.cantidad = (insumo.cantidad - cantidad)
        elif tipo in ("ENTRADA", "AJUSTE"):
            insumo.cantidad = (insumo.cantidad + cantidad)
        else:
            raise ValidationError({"tipo": "Tipo inválido"})

        insumo.save(update_fields=["cantidad"])

        mov = InsumoMovimiento.objects.create(
            insumo=insumo,
            tercero=tercero,
            bodega=bodega or insumo.bodega,
            tipo=tipo,
            cantidad=cantidad,
            unidad_medida=getattr(insumo, "unidad_medida", "") or "",
            costo_unitario=costo_unitario,
            total=total,
            saldo_resultante=insumo.cantidad,
            factura=factura or getattr(insumo, "factura", "") or "",
            observacion=observacion or "",
            nota_ensamble=nota_ensamble,
        )

    return mov

def _parse_decimal(v, field):
    try:
        if v is None: 
            return None
        s = str(v).strip()
        if not s: 
            return None
        # Normalizar coma a punto
        s = s.replace(",", ".")
        return Decimal(s).quantize(Decimal("0.001"))
    except Exception:
        raise ValidationError({field: f"Valor inválido: {v}"})


def _parse_date(v, field):
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if hasattr(v, "year") and hasattr(v, "month") and hasattr(v, "day"):
        return v
    # strings
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    raise ValidationError({field: f"Fecha inválida: {v}. Formatos: YYYY-MM-DD o DD/MM/YYYY"})

class DebugValidationMixin:
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        try:
            serializer.is_valid(raise_exception=True)
        except ValidationError as e:
            return Response(
                {
                    "detail": "Error de validación (DEBUG).",
                    "errors": e.detail,
                    "received": request.data,
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        try:
            serializer.is_valid(raise_exception=True)
        except ValidationError as e:
            return Response(
                {
                    "detail": "Error de validación (DEBUG).",
                    "errors": e.detail,
                    "received": request.data,
                },
                status=status.HTTP_400_BAD_REQUEST
            )
        self.perform_update(serializer)
        return Response(serializer.data, status=status.HTTP_200_OK)


from .services.inventory_service import InventoryService

class NotaEnsambleViewSet(viewsets.ModelViewSet):
    queryset = (
        NotaEnsamble.objects
        .prefetch_related("detalles", "insumos")  # ✅ importante
        .select_related("bodega", "tercero")
        .order_by("-id")
    )
    serializer_class = NotaEnsambleSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = NotaEnsambleFilter
    search_fields = [
        "observaciones", 
        "tercero__nombre", 
        "detalles__producto__nombre", 
        "detalles__producto__codigo_sku"
    ]
    ordering_fields = ["id", "fecha_elaboracion", "creado_en"]

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        """
        Delegamos la creación al InventoryService
        """
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        nota = InventoryService.create_assembly_note(serializer, serializer.validated_data)
        
        return Response(self.get_serializer(nota).data, status=status.HTTP_201_CREATED)

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        """
        Delegamos la actualización segura al InventoryService
        """
        nota = self.get_object()
        data = request.data

        # 1. Identificar cambios críticos (igual que antes para validaciones rápidas)
        has_detalles_input = "detalles_input" in data
        has_insumos_input = "insumos_input" in data
        has_bodega_change = "bodega_id" in data and int(data["bodega_id"]) != nota.bodega_id

        is_critical_change = has_detalles_input or has_insumos_input or has_bodega_change

        if is_critical_change:
            # Validaciones de seguridad (Bloqueos si ya se usó la nota)
            details_in_other_bodega = nota.detalles.filter(bodega_actual__isnull=False).exclude(bodega_actual_id=nota.bodega_id).exists()
            if details_in_other_bodega:
                raise ValidationError({
                    "detail": "Esta nota tiene productos trasladados a otra bodega. Revierta traslados.",
                    "code": "NOTA_CON_TRASLADOS"
                })

            if NotaSalidaAfectacionStock.objects.filter(detalle_stock__nota=nota).exists():
                raise ValidationError({
                    "detail": "Esta nota tiene productos vendidos. Elimine las ventas asociadas.",
                    "code": "NOTA_CON_SALIDAS"
                })

            if TrasladoProducto.objects.filter(detalle__nota=nota).exists():
                raise ValidationError({
                    "detail": "Esta nota tiene historial de traslados. Revierta traslados.",
                    "code": "NOTA_CON_HISTORIAL_TRASLADOS"
                })

        # 2. Delegar al servicio
        serializer = self.get_serializer(nota, data=data, partial=True)
        serializer.is_valid(raise_exception=True)
        
        updated_nota = InventoryService.update_assembly_note(nota, serializer, serializer.validated_data)

        return Response(self.get_serializer(updated_nota).data, status=status.HTTP_200_OK)

    @transaction.atomic
    def destroy(self, request, *args, **kwargs):
        """
        Bloquea eliminación si hay registrados movimientos que dependan de esta nota.
        """
        nota = self.get_object()

        # Validaciones de seguridad para eliminación
        # 🚫 Solo bloquear si bodega_actual es No Nulo Y Distinto a la bodega original
        details_in_other_bodega = nota.detalles.filter(bodega_actual__isnull=False).exclude(bodega_actual_id=nota.bodega_id).exists()
        if details_in_other_bodega:
            raise ValidationError({"detail": "No se puede eliminar: tiene productos en otras bodegas."})

        if NotaSalidaAfectacionStock.objects.filter(detalle_stock__nota=nota).exists():
            raise ValidationError({"detail": "No se puede eliminar: ya tiene productos vendidos."})
        
        if TrasladoProducto.objects.filter(detalle__nota=nota).exists():
            raise ValidationError({"detail": "No se puede eliminar: tiene historial de traslados."})

        # Revertir stock antes de borrar, dejando una traza en el historial
        obs_del = f"Eliminación nota #{nota.id}"
        self._aplicar_detalles(nota, list(nota.detalles.all()), signo=Decimal("-1"), observacion_p=obs_del)
        self._aplicar_insumos_manuales(nota, signo=Decimal("-1"), observacion_p=obs_del)

        return super().destroy(request, *args, **kwargs)
  


class DebugValidationMixin:
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        try:
            serializer.is_valid(raise_exception=True)
        except ValidationError as e:
            return Response(
                {"detail": "Error de validación (DEBUG).", "errors": e.detail, "received": request.data},
                status=status.HTTP_400_BAD_REQUEST
            )
        self.perform_create(serializer)
        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        try:
            serializer.is_valid(raise_exception=True)
        except ValidationError as e:
            return Response(
                {"detail": "Error de validación (DEBUG).", "errors": e.detail, "received": request.data},
                status=status.HTTP_400_BAD_REQUEST
            )
        self.perform_update(serializer)
        return Response(serializer.data, status=status.HTTP_200_OK)


class ProveedorViewSet(viewsets.ModelViewSet):
    # Ordenar por activos primero
    queryset = Proveedor.objects.all().order_by("-es_activo", "id")
    serializer_class = ProveedorSerializer
    ordering_fields = ["nombre", "es_activo"]

    def perform_destroy(self, instance):
        # Soft delete
        instance.es_activo = False
        instance.save(update_fields=["es_activo"])


class BodegaViewSet(viewsets.ModelViewSet):
    queryset = Bodega.objects.all().order_by("-es_activo", "nombre")
    serializer_class = BodegaSerializer
    filterset_fields = ["ubicacion"]
    search_fields = ["nombre", "codigo"]
    ordering_fields = ["nombre", "codigo", "es_activo"]

    def perform_destroy(self, instance):
        instance.es_activo = False
        instance.save(update_fields=["es_activo"])

    def get_queryset(self):
        return (
            Bodega.objects
            .annotate(
                insumos_count=Count("insumos", distinct=True),

                # Cuenta productos distintos en los detalles cuya bodega efectiva sea esta bodega
                productos_count=Count(
                    "productos_detalle__producto",
                    filter=Q(productos_detalle__cantidad__gt=0),
                    distinct=True
                )
            )
            .order_by("nombre")
        )   

    @action(detail=True, methods=["get"], url_path="stock-terminado")
    def stock_terminado(self, request, pk=None):
        """
        Devuelve el stock de producto terminado en esta bodega, agrupado por SKU + talla.
        Soporta filtro opcional: ?sku=CAM-005
        """
        bodega = self.get_object()
        sku = (request.query_params.get("sku") or "").strip()

        qs = (
            NotaEnsambleDetalle.objects
            .filter(cantidad__gt=0)
            .filter(
                Q(bodega_actual=bodega) |
                Q(bodega_actual__isnull=True, nota__bodega=bodega)
            )
        )

        if sku:
            qs = qs.filter(producto__codigo_sku=sku)

        data = (
            qs.values("producto__codigo_sku", "producto__nombre", "talla__nombre")
              .annotate(cantidad=Sum("cantidad_disponible"))
              .order_by("producto__codigo_sku", "talla__nombre")
        )

        # Normaliza keys a lo que el front consume fácil
        result = [
            {
                "producto_id": row["producto__codigo_sku"],
                "producto_nombre": row["producto__nombre"],
                "talla": row["talla__nombre"] or "",
                "cantidad": str(row["cantidad"] or 0),
                "bodega_id": bodega.id,
                "bodega_nombre": bodega.nombre,
            }
            for row in data
        ]

        return Response(result)

    @action(detail=True, methods=["get"], url_path="contenido")
    def contenido(self, request, pk=None):
        bodega = self.get_object()

        # ✅ Insumos de la bodega (tu Insumo tiene FK bodega) :contentReference[oaicite:1]{index=1}
        insumos_qs = (
            Insumo.objects
            .filter(bodega=bodega)
            .order_by("nombre")
        )

        insumos = []
        for i in insumos_qs:
            stock = i.cantidad or 0
            cu = i.costo_unitario or 0
            total = (stock * cu) if stock and cu else 0
            insumos.append({
                "codigo": i.codigo,
                "nombre": i.nombre,
                "stock_actual": str(stock),
                "costo_unitario": str(cu),
                "valor_total": str(total),
            })

        # ✅ Productos producidos en esa bodega:
        # NotaEnsamble tiene FK bodega, y NotaEnsambleDetalle tiene producto + cantidad :contentReference[oaicite:2]{index=2}
        productos_qs = (
            NotaEnsambleDetalle.objects
            .select_related("producto", "nota", "bodega_actual")
            .annotate(bodega_efectiva=Coalesce("bodega_actual_id", "nota__bodega_id"))
            .filter(bodega_efectiva=bodega.id)
            .values("producto__codigo_sku", "producto__nombre")
            .annotate(total_producido=Sum("cantidad"), stock_actual=Sum("cantidad_disponible"))
            .order_by("producto__nombre")
        )

        productos = []
        for p in productos_qs:
            productos.append({
                "codigo": p["producto__codigo_sku"],
                "nombre": p["producto__nombre"],
                "total_producido": str(p["total_producido"] or 0),
                "stock_actual": str(p["stock_actual"] or 0),
            })

        return Response({
            "bodega": {"id": bodega.id, "codigo": bodega.codigo, "nombre": bodega.nombre},
            "insumos": insumos,
            "productos": productos,
        })


class TerceroViewSet(viewsets.ModelViewSet):
    queryset = Tercero.objects.all().order_by("-es_activo", "codigo")
    serializer_class = TerceroSerializer
    search_fields = ["nombre", "codigo"]
    ordering_fields = ["nombre", "es_activo"]

    def perform_destroy(self, instance):
        instance.es_activo = False
        instance.save(update_fields=["es_activo"])


class ImpuestoViewSet(viewsets.ModelViewSet):
    queryset = Impuesto.objects.all().order_by("-es_activo", "nombre")
    serializer_class = ImpuestoSerializer

    def perform_destroy(self, instance):
        instance.es_activo = False
        instance.save(update_fields=["es_activo"])


class ProductoViewSet(viewsets.ModelViewSet):
    queryset = (
        Producto.objects
        .select_related("tercero")
        .prefetch_related("impuestos", "precios", "datos_adicionales")
        .annotate(
            tiene_bajo_stock=Case(
                When(datos_adicionales__stock__lt=F("datos_adicionales__stock_minimo"), then=Value(1)),
                default=Value(2),
                output_field=IntegerField(),
            )
        )
        .order_by("-es_activo", "tiene_bajo_stock", "-creado_en")
    )
    serializer_class = ProductoSerializer
    filterset_class = ProductoFilter
    search_fields = ["nombre", "codigo_sku", "codigo_barras"]
    ordering_fields = ["nombre", "creado_en", "es_activo"]

    def perform_destroy(self, instance):
        instance.es_activo = False
        instance.save(update_fields=["es_activo"])

    @action(detail=True, methods=["get"], url_path="stock-por-talla")
    def stock_por_talla(self, request, pk=None):
        producto = self.get_object()

        bodega_id = request.query_params.get("bodega_id")
        if bodega_id is not None and not str(bodega_id).isdigit():
            raise ValidationError({"bodega_id": "Debe ser un entero."})

        qs = (
            NotaEnsambleDetalle.objects
            .select_related("talla", "nota", "bodega_actual")
            .filter(producto=producto)
            .annotate(bodega_efectiva=Coalesce("bodega_actual_id", "nota__bodega_id"))
        )

        # ✅ si viene bodega_id, filtra por esa bodega
        if bodega_id:
            qs = qs.filter(bodega_efectiva=int(bodega_id))

        items = (
            qs.values("talla__nombre")
            .annotate(cantidad=Sum("cantidad_disponible"))
            .order_by("talla__nombre")
        )

        return Response({
            "producto": {
                "codigo": producto.codigo_sku,
                "nombre": producto.nombre,
            },
            "bodega_id": int(bodega_id) if bodega_id else None,
            "items": [
                {
                    "codigo": producto.codigo_sku,
                    "nombre": producto.nombre,
                    "talla_id": it["talla__nombre"],
                    "talla": it["talla__nombre"] or "Sin talla",
                    "cantidad": str(it["cantidad"] or 0),
                }
                for it in items
            ],
        })


class PrecioProductoViewSet(viewsets.ModelViewSet):
    queryset = PrecioProducto.objects.select_related("producto").order_by("-id")
    serializer_class = ProductoPrecioWriteSerializer


class DatosAdicionalesProductoViewSet(DebugValidationMixin, viewsets.ModelViewSet):
    queryset = DatosAdicionalesProducto.objects.select_related("producto").order_by("-id")
    serializer_class = DatosAdicionalesWriteSerializer


class InsumoViewSet(viewsets.ModelViewSet):
    # Ordenar primero por activos vs inactivos, luego por bajo stock, luego nombre
    queryset = (
        Insumo.objects.select_related("bodega", "proveedor", "tercero")
        .annotate(
            tiene_bajo_stock=Case(
                When(cantidad__lt=F("stock_minimo"), then=Value(1)),
                default=Value(2),
                output_field=IntegerField(),
            )
        )
        .order_by("-es_activo", "tiene_bajo_stock", "nombre")
    )
    serializer_class = InsumoSerializer
    filterset_class = InsumoFilter
    search_fields = ["nombre", "codigo", "referencia", "observacion"]
    ordering_fields = ["nombre", "cantidad", "costo_unitario", "creado_en", "es_activo"]

    def perform_destroy(self, instance):
        # Soft delete: Marcar como inactivo en lugar de borrar
        instance.es_activo = False
        instance.save(update_fields=["es_activo"])

    def perform_update(self, serializer):
        insumo = serializer.save()
        # Registrar EDICION en historial (sin afectar stock)
        tercero = insumo.tercero
        registrar_movimiento_sin_afectar_stock(
            insumo=insumo,
            tercero=tercero,  # puede ser None, pero la función lo maneja o fallará si el modelo lo exige (el modelo InsumoMovimiento.tercero NO es null=True, ojo)
            # Pero Insumo.tercero es null=True? models.py dice: tercero = models.ForeignKey("Tercero", ..., null=True, blank=True)
            # InsumoMovimiento.tercero NO es null=True.
            # Necesitamos un tercero para el historial. Si el insumo no tiene tercero, ¿qué hacemos?
            # Usar uno por defecto o validar.
            # REVISAR models.py: InsumoMovimiento line 424: tercero = models.ForeignKey("Tercero", on_delete=models.PROTECT) -> NO NULL.
            # PROBABLE BUG IF INSUMO HAS NO TERCERO.
            # Veamos perform_create:
            #   tercero = insumo.tercero
            #   if tercero:
            #       registrar_...
            # O sea solo registra si tiene tercero. Mantengamos esa lógica.
            tipo="EDICION",
            cantidad=Decimal("0.000"),
            costo_unitario=insumo.costo_unitario,
            bodega=insumo.bodega,
            factura=getattr(insumo, "factura", "") or "",
            observacion="Actualización de datos",
        ) if insumo.tercero else None

    def perform_create(self, serializer):
        insumo = serializer.save()

        # Registra "CREACION" en historial (sin afectar stock) usando el tercero del insumo.
        # Esto evita doble conteo porque el insumo ya quedó con cantidad guardada.
        tercero = insumo.tercero
        if tercero:
            registrar_movimiento_sin_afectar_stock(
                insumo=insumo,
                tercero=tercero,
                tipo="CREACION",
                cantidad=insumo.cantidad,
                costo_unitario=insumo.costo_unitario,
                bodega=insumo.bodega,
                factura=getattr(insumo, "factura", "") or "",
                observacion="Creación de insumo",
            )

    @action(detail=True, methods=["get"], url_path="movimientos")
    def movimientos(self, request, pk=None):
        """
        GET /insumos/{codigo}/movimientos/?page=... (paginación la da DRF si la tienes global)
        """
        insumo = self.get_object()
        qs = InsumoMovimiento.objects.select_related("insumo", "tercero", "bodega").filter(insumo=insumo)

        tipo = request.query_params.get("tipo")
        if tipo:
            qs = qs.filter(tipo=tipo)

        tercero_id = request.query_params.get("tercero_id")
        if tercero_id:
            qs = qs.filter(tercero_id=tercero_id)

        bodega_id = request.query_params.get("bodega_id")
        if bodega_id:
            qs = qs.filter(bodega_id=bodega_id)

        # paginación DRF
        page = self.paginate_queryset(qs.order_by("-fecha", "-id"))
        if page is not None:
            ser = InsumoMovimientoSerializer(page, many=True)
            return self.get_paginated_response(ser.data)

        return Response(InsumoMovimientoSerializer(qs, many=True).data)

    @action(detail=True, methods=["get"], url_path="stock_por_bodega")
    def stock_por_bodega(self, request, pk=None):
        """
        Retorna el stock calculado por bodega basado en el historial de movimientos.
        """
        insumo = self.get_object()
        
        # Agrupar por bodega y sumarizar
        # ENTRADA, AJUSTE, CREACION suman (o ajustan)
        # SALIDA, CONSUMO_ENSAMBLE restan
        
        from django.db.models import Sum, Case, When, F, DecimalField, Value
        
        # Primero, obtenemos todas las bodegas que han tenido movimiento con este insumo
        movs = InsumoMovimiento.objects.filter(insumo=insumo).values("bodega", "bodega__nombre")
        
        results = movs.annotate(
            total_entradas=Sum(
                Case(
                    When(tipo__in=["CREACION", "ENTRADA", "AJUSTE"], then=F("cantidad")),
                    default=0,
                    output_field=DecimalField()
                )
            ),
            total_salidas=Sum(
                Case(
                    When(tipo__in=["SALIDA", "CONSUMO_ENSAMBLE"], then=F("cantidad")),
                    default=0,
                    output_field=DecimalField()
                )
            )
        ).order_by("bodega__nombre")
        
        # Procesar resultados
        data = []
        for r in results:
            bodega_nombre = r["bodega__nombre"]
            if not bodega_nombre:
                bodega_nombre = "Sin Bodega"
                
            stock = (r["total_entradas"] or 0) - (r["total_salidas"] or 0)
            
            # Solo mostrar si hay algo relevante (o si es 0 pero hubo movs)
            data.append({
                "bodega_id": r["bodega"],
                "bodega_nombre": bodega_nombre,
                "stock": stock
            })
            
        return Response(data)
    @action(detail=True, methods=["post"], url_path="movimiento")
    def movimiento(self, request, pk=None):
        """
        POST /insumos/{codigo}/movimiento/
        Body: { tipo: ENTRADA|SALIDA|AJUSTE, tercero_id, cantidad, costo_unitario?, bodega_id?, factura?, observacion? }
        """
        insumo = self.get_object()
        
        inp = InsumoMovimientoInputSerializer(data=request.data)
        inp.is_valid(raise_exception=True)
        data = inp.validated_data

        tipo = data["tipo"]
        tercero = Tercero.objects.get(id=data["tercero_id"])
        bodega = None
        if data.get("bodega_id"):
            bodega = Bodega.objects.get(id=data["bodega_id"])

        # ENTRADA/SALIDA/AJUSTE afectan stock
        mov = aplicar_movimiento_insumo(
            insumo=insumo,
            tercero=tercero,
            tipo=tipo,
            cantidad=data["cantidad"],
            costo_unitario=data.get("costo_unitario", None),
            bodega=bodega,
            factura=data.get("factura", ""),
            observacion=data.get("observacion", ""),
        )

        return Response(InsumoMovimientoSerializer(mov).data, status=status.HTTP_201_CREATED)



class TallaViewSet(viewsets.ModelViewSet):
    queryset = Talla.objects.all().order_by("-es_activo", "nombre")
    serializer_class = TallaSerializer
    lookup_field = "nombre"
    search_fields = ["nombre"]
    ordering_fields = ["nombre", "es_activo"]

    def perform_destroy(self, instance):
        instance.es_activo = False
        instance.save(update_fields=["es_activo"])
    # (Dejo tu lógica específica fuera por ahora porque en tu código original
    #  estabas llamando self._get_datos_adicionales aquí y eso NO existe en TallaViewSet.
    #  Si esa lógica era para otra cosa, me dices y la reubicamos bien.)


class ProductoInsumoViewSet(viewsets.ModelViewSet):
    queryset = ProductoInsumo.objects.select_related("producto", "insumo").all()
    serializer_class = ProductoInsumoSerializer

class TrasladoProductoViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Historial de traslados (GET) y endpoint de ejecutar traslado (POST /traslados-producto/ejecutar/)
    """
    queryset = (
        TrasladoProducto.objects
        .select_related("tercero", "bodega_origen", "bodega_destino", "producto", "talla", "detalle")
        .order_by("-id")
    )
    serializer_class = TrasladoProductoSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        bodega_id = self.request.query_params.get("bodega_id")
        if bodega_id and str(bodega_id).isdigit():
            bodega_id = int(bodega_id)
            qs = qs.filter(
                Q(bodega_origen_id=bodega_id) |
                Q(bodega_destino_id=bodega_id)
            )
        return qs

    @action(detail=False, methods=["post"], url_path="ejecutar-masivo")
    @transaction.atomic
    def ejecutar_masivo(self, request):
        """
        Payload:
        {
          "tercero_id": 1,
          "bodega_origen_id": 2,
          "bodega_destino_id": 3,
          "items": [
             { "producto_id": "SKU", "talla_id": 5, "cantidad": "2" },
             ...
          ]
        }
        """
        data = request.data
        tercero_id = data.get("tercero_id")
        origen_id = data.get("bodega_origen_id")
        destino_id = data.get("bodega_destino_id")
        items = data.get("items", [])

        if not items:
            raise ValidationError({"items": "La lista de items está vacía."})

        if origen_id == destino_id:
            raise ValidationError({"bodega_destino_id": "Destino debe ser diferente a origen."})

        tercero = get_object_or_404(Tercero, pk=tercero_id)
        b_origen = get_object_or_404(Bodega, pk=origen_id)
        b_destino = get_object_or_404(Bodega, pk=destino_id)

        ok_count = 0
        
        for item in items:
            sku = item.get("producto_id")
            talla_id = item.get("talla_id")
            cantidad_str = item.get("cantidad")
            
            try:
                producto = Producto.objects.get(codigo_sku=sku)
            except Producto.DoesNotExist:
                raise ValidationError(f"Producto {sku} no existe.")

            talla = None
            if talla_id:
                talla = Talla.objects.get(pk=talla_id)

            cantidad = _d(cantidad_str)
            if cantidad <= 0:
                raise ValidationError(f"Cantidad inválida para {sku}.")

            # --- Lógica de traslado (reutilizada de 'ejecutar') ---
            qs = (
                NotaEnsambleDetalle.objects
                .select_related("nota", "nota__bodega", "bodega_actual")
                .annotate(bodega_efectiva=Coalesce("bodega_actual_id", "nota__bodega_id"))
                .filter(producto=producto)
                .filter(bodega_efectiva=b_origen.id)
                .order_by("nota__fecha_elaboracion", "id")
            )

            if talla is None:
                qs = qs.filter(talla__isnull=True)
            else:
                qs = qs.filter(talla=talla)
            
            # Bloquear filas para evitar race conditions
            qs = qs.select_for_update()

            disponible_total = sum(_d(x.cantidad_disponible) for x in qs)
            if disponible_total < cantidad:
                talla_nombre = talla.nombre if talla else "Única"
                raise ValidationError({
                    "stock_insuficiente": {
                        "producto": f"{producto.nombre} ({talla_nombre})",
                        "disponible": str(disponible_total),
                        "requerido": str(cantidad),
                        "faltante": str(cantidad - disponible_total),
                    }
                })

            restante = cantidad
            for det in qs:
                if restante <= 0: 
                    break
                
                disponible_det = _d(det.cantidad_disponible)
                mover = min(disponible_det, restante)
                if mover <= 0:
                    continue
                
                # 1. Restar origen
                det.cantidad_disponible = disponible_det - mover
                det.save(update_fields=["cantidad_disponible"])

                # 2. Sumar destino
                dest_det, _created = NotaEnsambleDetalle.objects.get_or_create(
                    nota=det.nota,
                    producto=det.producto,
                    talla=det.talla,
                    bodega_actual=b_destino,
                    defaults={"cantidad": Decimal("0"), "cantidad_disponible": Decimal("0")}
                )
                dest_det.cantidad_disponible = _d(dest_det.cantidad_disponible) + mover
                dest_det.save(update_fields=["cantidad_disponible"])

                # 3. Historial
                TrasladoProducto.objects.create(
                    tercero=tercero,
                    bodega_origen=b_origen,
                    bodega_destino=b_destino,
                    producto=producto,
                    talla=talla,
                    cantidad=mover,
                    detalle=det
                )
                restante -= mover
            
            ok_count += 1

        return Response({"ok": True, "items_movidos": ok_count}, status=status.HTTP_200_OK)

    @action(detail=False, methods=["post"], url_path="ejecutar")
    @transaction.atomic
    def ejecutar(self, request):
        """
        Payload:
        {
          "tercero_id": 1,
          "bodega_origen_id": 2,
          "bodega_destino_id": 3,
          "producto_id": "SKU-001",
          "talla_id": 5,          // opcional
          "cantidad": "2.000"
        }
        """
        ser = TrasladoProductoSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        v = ser.validated_data

        tercero = v["tercero"]
        b_origen = v["bodega_origen"]
        b_destino = v["bodega_destino"]
        producto = v["producto"]
        talla = v.get("talla", None)
        cantidad = _d(v["cantidad"])

        if b_origen.id == b_destino.id:
            raise ValidationError({"bodega_destino_id": "La bodega destino debe ser diferente a la bodega origen."})

        if cantidad <= 0:
            raise ValidationError({"cantidad": "Debe ser mayor que 0."})

        # Buscar detalles disponibles en la bodega origen (por bodega_actual; si null => nota.bodega)
        qs = (
            NotaEnsambleDetalle.objects
            .select_related("nota", "nota__bodega", "bodega_actual")
            .annotate(bodega_efectiva=Coalesce("bodega_actual_id", "nota__bodega_id"))
            .filter(producto=producto)
            .filter(bodega_efectiva=b_origen.id)
            .order_by("nota__fecha_elaboracion", "id")
        )

        if talla is None:
            qs = qs.filter(talla__isnull=True)
        else:
            qs = qs.filter(talla=talla)

        disponible_total = sum(_d(x.cantidad_disponible) for x in qs)
        if disponible_total < cantidad:
            raise ValidationError({
                "stock_insuficiente": {
                    "disponible": str(disponible_total),
                    "requerido": str(cantidad),
                    "faltante": str(cantidad - disponible_total),
                }
            })

        restante = cantidad

        for det in qs:
            if restante <= 0:
                break

            disponible_det = _d(det.cantidad_disponible)
            mover = min(disponible_det, restante)
            if mover <= 0:
                continue

            # 1) restar en origen (PERSISTIR)
            nuevo_origen = disponible_det - mover
            if nuevo_origen < 0:
                raise ValidationError("Error interno: cantidad disponible negativa tras traslado.")

            det.cantidad_disponible = nuevo_origen
            det.save(update_fields=["cantidad_disponible"])  # ✅ CLAVE: guardar la resta

            # 2) sumar/crear en destino manteniendo MISMA nota
            dest_det, _created = NotaEnsambleDetalle.objects.get_or_create(
                nota=det.nota,
                producto=det.producto,
                talla=det.talla,
                bodega_actual=b_destino,
                defaults={"cantidad": Decimal("0"), "cantidad_disponible": Decimal("0")}
            )
            dest_det.cantidad_disponible = _d(dest_det.cantidad_disponible) + mover
            dest_det.save(update_fields=["cantidad_disponible"])

            # 3) historial
            TrasladoProducto.objects.create(
                tercero=tercero,
                bodega_origen=b_origen,
                bodega_destino=b_destino,
                producto=producto,
                talla=talla,
                cantidad=mover,
                detalle=det
            )

            restante -= mover


        return Response({"ok": True, "cantidad_movida": str(cantidad)}, status=status.HTTP_200_OK)

class NotaSalidaProductoViewSet(viewsets.ModelViewSet):
    queryset = NotaSalidaProducto.objects.all().prefetch_related("detalles", "detalles__afectaciones", "bodega", "tercero")
    serializer_class = NotaSalidaProductoSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = NotaSalidaProductoFilter
    search_fields = ["numero", "observacion", "detalles__producto__nombre", "detalles__producto__codigo_sku"]
    ordering_fields = ["id", "fecha", "creado_en"]

    @transaction.atomic
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        # Revertir stock: devolver a las NotaEnsambleDetalle originales y al global
        for detalle in instance.detalles.all():
            # 1. Devolver a la bodega (FIFO afectaciones)
            for afectacion in detalle.afectaciones.all():
                stock_row = afectacion.detalle_stock
                stock_row.cantidad_disponible = (stock_row.cantidad_disponible + afectacion.cantidad)
                stock_row.save(update_fields=["cantidad_disponible"])
            
            # 2. Devolver al global
            datos = DatosAdicionalesProducto.objects.filter(producto=detalle.producto).first()
            if datos:
                datos.stock = (datos.stock + detalle.cantidad)
                datos.save(update_fields=["stock"])
        
        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=["get"], url_path="pdf")
    def pdf(self, request, pk=None):
        salida = get_object_or_404(
            NotaSalidaProducto.objects.prefetch_related("detalles", "detalles__afectaciones"),
            pk=pk
        )

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="{salida.numero}.pdf"'

        c = canvas.Canvas(response, pagesize=letter)
        width, height = letter

        y = height - 50
        c.setFont("Helvetica-Bold", 14)
        c.drawString(50, y, f"NOTA DE SALIDA - {salida.numero}")
        y -= 20

        c.setFont("Helvetica", 10)
        c.drawString(50, y, f"Fecha: {salida.fecha}")
        y -= 14
        c.drawString(50, y, f"Bodega: {salida.bodega.codigo} - {salida.bodega.nombre}")
        y -= 14
        tercero_txt = f"{salida.tercero.codigo} - {salida.tercero.nombre}" if salida.tercero else "-"
        c.drawString(50, y, f"Tercero: {tercero_txt}")
        y -= 14

        if salida.observacion:
            c.drawString(50, y, f"Observación: {salida.observacion[:120]}")
            y -= 14

        y -= 10
        c.setFont("Helvetica-Bold", 10)
        c.drawString(50, y, "Detalle")
        y -= 14

        c.setFont("Helvetica-Bold", 9)
        c.drawString(50, y, "SKU")
        c.drawString(120, y, "Producto")
        c.drawString(320, y, "Talla")
        c.drawString(370, y, "Cantidad")
        c.drawString(450, y, "Costo U.")
        c.drawString(520, y, "Total")
        y -= 10
        c.line(50, y, 560, y)
        y -= 12

        c.setFont("Helvetica", 9)
        for d in salida.detalles.all():
            if y < 80:
                c.showPage()
                y = height - 50
                c.setFont("Helvetica-Bold", 9)
                c.drawString(50, y, "SKU")
                c.drawString(120, y, "Producto")
                c.drawString(320, y, "Talla")
                c.drawString(370, y, "Cantidad")
                c.drawString(450, y, "Costo U.")
                c.drawString(520, y, "Total")
                y -= 10
                c.line(50, y, 560, y)
                y -= 12
                c.setFont("Helvetica", 9)

            sku = d.producto.codigo_sku
            nombre = d.producto.nombre[:30]
            talla = d.talla or "-"
            cantidad = str(d.cantidad)
            cu = str(d.costo_unitario) if d.costo_unitario is not None else "-"
            total = str(d.total) if d.total is not None else "-"

            c.drawString(50, y, sku)
            c.drawString(120, y, nombre)
            c.drawString(320, y, talla)
            c.drawRightString(420, y, cantidad)
            c.drawRightString(505, y, cu)
            c.drawRightString(560, y, total)
            y -= 14

        c.showPage()
        c.save()
        return response

class InsumoMovimientoViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Kardex global:
    GET /insumo-movimientos/?insumo=INS-001&tipo=ENTRADA&tercero_id=1&bodega_id=2
    """
    queryset = InsumoMovimiento.objects.select_related("insumo", "tercero", "bodega").all()
    serializer_class = InsumoMovimientoSerializer

    def get_queryset(self):
        qs = super().get_queryset()

        insumo_codigo = self.request.query_params.get("insumo")
        if insumo_codigo:
            qs = qs.filter(insumo_id=insumo_codigo)

        tipo = self.request.query_params.get("tipo")
        if tipo:
            qs = qs.filter(tipo=tipo)

        tercero_id = self.request.query_params.get("tercero_id")
        if tercero_id:
            qs = qs.filter(tercero_id=tercero_id)

        bodega_id = self.request.query_params.get("bodega_id")
        if bodega_id:
            qs = qs.filter(bodega_id=bodega_id)

        return qs.order_by("-fecha", "-id")

class ExcelImportViewSet(viewsets.ViewSet):
    """
    Endpoints:
      GET  /api/excel/plantilla-insumos/
      POST /api/excel/importar-insumos/         (multipart: file)
      GET  /api/excel/plantilla-terminado/
      POST /api/excel/importar-terminado/       (multipart: file)

      GET  /api/excel/kardex-terminado/?sku=...&bodega_id=...&tercero_id=...
    """

    @action(detail=False, methods=["get"], url_path="plantilla-insumos", renderer_classes=[XLSXRenderer])
    def plantilla_insumos(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "PlantillaInsumos"

        # --- Hoja Principal ---
        headers = [
            "Codigo Producto", "Descripción", "Cantidad Entrada (Stock)",
            "Costo Unitario", "Marca (Proveedor)", "Color", "Factura", 
            "Bodega", "Tercero", "Unidad Medida"
        ]
        
        # Estilos
        bold_white = Font(bold=True, color="FFFFFF")
        dark_blue_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Escribir headers en fila 2, col 2 (B2)
        start_row = 2
        start_col = 2
        
        for idx, h in enumerate(headers):
            cell = ws.cell(row=start_row, column=start_col + idx, value=h)
            cell.font = bold_white
            cell.fill = dark_blue_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            
            # Ajustar ancho estimado
            ws.column_dimensions[chr(65 + start_col + idx - 1)].width = len(h) + 5
            
        # Ejemplo de datos
        ws.cell(row=start_row + 1, column=start_col, value="INS-001").alignment = Alignment(horizontal="left")
        ws.cell(row=start_row + 1, column=start_col + 1, value="Ej: Botón Rojo 5mm")
        ws.cell(row=start_row + 1, column=start_col + 2, value=100)
        ws.cell(row=start_row + 1, column=start_col + 3, value=50.00)
        ws.cell(row=start_row + 1, column=start_col + 4, value="Proveedor ACME")
        ws.cell(row=start_row + 1, column=start_col + 5, value="Rojo")
        ws.cell(row=start_row + 1, column=start_col + 6, value="FAC-1234")
        ws.cell(row=start_row + 1, column=start_col + 7, value="Bodega Principal")
        ws.cell(row=start_row + 1, column=start_col + 8, value="Proveedor Genérico (ID: 1)")
        ws.cell(row=start_row + 1, column=start_col + 9, value="UN")

        # --- Hoja Referencias (Ayuda) ---
        ws_ref = wb.create_sheet("Referencias")
        ws_ref_headers = ["Unidad (Abreviatura)", "Descripción", "Ejemplo"]
        
        for idx, h in enumerate(ws_ref_headers):
            cell = ws_ref.cell(row=1, column=idx+1, value=h)
            cell.font = bold_white
            cell.fill = dark_blue_fill
            cell.border = border
            ws_ref.column_dimensions[chr(65 + idx)].width = 25

        referencias = [
            ("UN", "Unidad / Pieza", "Botones, Cremalleras"),
            ("M", "Metros", "Tela, Elástico"),
            ("CM", "Centímetros", "Cinta"),
            ("KG", "Kilogramos", "Relleno"),
            ("GR", "Gramos", "Hilo"),
            ("L", "Litros", "Tintas"),
            ("PAR", "Par", "Zapatos, Guantes"),
            ("SET", "Juego / Kit", "Kit de reparación"),
            ("CAJA", "Caja", "Caja de hilos"),
            ("ROLLO", "Rollo", "Rollo de tela"),
        ]

        for r_idx, row_data in enumerate(referencias, start=2):
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws_ref.cell(row=r_idx, column=c_idx, value=val)
                cell.border = border

        # Agregar nota en la hoja principal apuntando a las referencias
        ws.merge_cells(start_row=start_row - 1, start_column=start_col, end_row=start_row - 1, end_column=start_col + len(headers) - 1)
        note_cell = ws.cell(row=start_row - 1, column=start_col, value="Nota: Para 'Unidad Medida' consulta la pestaña 'Referencias' para ver las abreviaturas recomendadas.")
        note_cell.font = Font(italic=True, color="555555", size=9)

        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()

        response = Response(content, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="plantilla_insumos_cliente.xlsx"'
        return response

    @action(detail=False, methods=["post"], url_path="importar-insumos")
    @transaction.atomic
    def importar_insumos(self, request):
        file = request.FILES.get("file")
        if not file:
            raise ValidationError({"file": "Debe enviar un archivo .xlsx en multipart/form-data con key 'file'."})

        # ✅ 1) Default Bodega / Tercero
        default_bodega_id = request.data.get("bodega_id")
        default_tercero_id = request.data.get("tercero_id")

        if getattr(file, "size", 0) == 0:
            raise ValidationError({"file": "El archivo llegó vacío (0 bytes). Revisa el FormData en el frontend."})

        # ✅ 2) Validar firma ZIP
        head = file.read(4)
        file.seek(0)
        if head[:2] != b"PK":
            raise ValidationError({"file": "El archivo no es un Excel válido (.xlsx)."})

        try:
            wb = load_workbook(filename=file, data_only=True)
        except Exception as e:
            raise ValidationError({"file": f"No se pudo leer el Excel: {str(e)}"})

        # Hoja activa
        ws = wb.active
        for sheet in wb.sheetnames:
            if "Insumos" in sheet or "Inventario" in sheet:
                ws = wb[sheet]
                break

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValidationError("El Excel está vacío.")

        # --- Detección inteligente de Header ---
        # Buscamos en las primeras 10 filas alguna que tenga palabras clave
        keywords = ["codigo", "descripcion", "producto", "stock", "cantidad", "marca", "bodega", "tercero"]
        
        def clean_header(h):
            if not h: return ""
            return str(h).lower().replace("ó", "o").replace("í", "i").replace("á", "a").replace("é", "e").replace("ú", "u").replace(".", "").strip()

        header_row_idx = -1
        raw_header = []
        
        for i, row in enumerate(rows[:10]):
            cleaned_row = [clean_header(c) for c in row]
            matches = sum(1 for cell in cleaned_row if any(k in cell for k in keywords))
            # Si encontramos al menos 2 palabras clave, asumimos que es el header
            if matches >= 2:
                header_row_idx = i
                raw_header = row
                break
        
        if header_row_idx == -1:
             # Fallback: intentar primera fila si no se encontró nada claro
             header_row_idx = 0
             raw_header = rows[0]

        # Mapa de alias (Cliente -> Backend)
        aliases = {
            "codigo producto": "codigo", "codigo": "codigo", "referencia": "codigo",
            "descripción": "nombre", "descripcion": "nombre", "producto": "nombre", "nombre": "nombre",
            "cantidad entrada (stock)": "cantidad_entrada", "stock actual": "cantidad_entrada", "stock": "cantidad_entrada", 
            "cantidad": "cantidad_entrada", "entradas": "cantidad_entrada", "cantidad_entrada": "cantidad_entrada",
            "marca (proveedor)": "proveedor_nombre", "marca": "proveedor_nombre", "fabricante": "proveedor_nombre", "proveedor": "proveedor_nombre",
            "unidad medida": "unidad_medida", "unidad_medida": "unidad_medida", "unidad": "unidad_medida", "medida": "unidad_medida", "um": "unidad_medida",
            "# factura": "factura", "factura": "factura",
            "costo unitario": "costo_unitario", "costo": "costo_unitario",
            "bodega": "bodega", "id_bodega": "bodega", "bodega_id": "bodega", "bodega*": "bodega",
            "tercero": "tercero", "id_tercero": "tercero", "tercero_id": "tercero", "tercero*": "tercero",
            "color": "color",
            "observacion": "observacion",
        }

        idx = {}
        # Normalizamos el header encontrado
        for i, h in enumerate(raw_header):
            cleaned = clean_header(h)
            if cleaned in aliases:
                key = aliases[cleaned]
                if key not in idx: idx[key] = i
            elif cleaned not in idx:
                idx[cleaned] = i # Fallback

        # Validar requeridos mínimos
        if "codigo" not in idx:
             # Generar mensaje amigable
             msg = f"No se encontró la columna 'Codigo Producto' en la fila de encadenados (fila {header_row_idx+1}). Cabeceras detectadas: {raw_header}"
             raise ValidationError(msg)

        ok = 0
        errores = []
        movimientos_creados = []

        # Caches para evitar DB hits masivos
        cache_bodegas = {b.nombre.lower(): b for b in Bodega.objects.all()} # nombre_lower -> obj
        cache_bodegas_id = {str(b.id): b for b in cache_bodegas.values()}   # str(id) -> obj
        cache_bodegas_cod = {b.codigo.lower(): b for b in cache_bodegas.values() if b.codigo} # codigo_lower -> obj
        
        cache_terceros = {t.nombre.lower(): t for t in Tercero.objects.all()}
        cache_terceros_id = {str(t.id): t for t in cache_terceros.values()}
        cache_terceros_cod = {t.codigo.lower(): t for t in cache_terceros.values() if t.codigo}
        
        cache_proveedores = {p.nombre.upper(): p for p in Proveedor.objects.all()}

        default_bodega_obj = None
        if default_bodega_id:
             default_bodega_obj = Bodega.objects.filter(pk=default_bodega_id).first()

        default_tercero_obj = None
        if default_tercero_id:
             default_tercero_obj = Tercero.objects.filter(pk=default_tercero_id).first()

        # Iterar desde la fila siguiente al header
        for i, r in enumerate(rows[header_row_idx+1:], start=header_row_idx+2):
            try:
                def get_val(key, default=None):
                    if key in idx and idx[key] < len(r):
                        val = r[idx[key]]
                        return val if val is not None else default
                    return default

                codigo = str(get_val("codigo", "")).strip()
                if not codigo: continue

                nombre = str(get_val("nombre", "")).strip()
                
                # --- Cantidad ---
                c_raw = get_val("cantidad_entrada", 0)
                cantidad_entrada = _parse_decimal(c_raw, "cantidad") or Decimal("0")

                # --- Costo ---
                costo_raw = get_val("costo_unitario", 0)
                val_costo = _parse_decimal(costo_raw, "costo") or Decimal("0")
                # Forzar 2 decimales para evitar error "más de 2 decimales"
                costo_unitario = val_costo.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

                # --- Resolver Bodega ---
                # Prioridad: 1. ID exacto, 2. Codigo exacto, 3. Nombre
                # Excel "0" suele ser string "0". Si codigo es "000", no machea directo.
                # Intento machear tal cual, y si es "0", intento "000".
                bodega_val = str(get_val("bodega", "")).strip()
                bodega_obj = default_bodega_obj
                
                if bodega_val:
                    b_look = bodega_val.lower()
                    # 1. ID
                    if bodega_val in cache_bodegas_id:
                        bodega_obj = cache_bodegas_id[bodega_val]
                    # 2. Codigo
                    elif b_look in cache_bodegas_cod:
                        bodega_obj = cache_bodegas_cod[b_look]
                    # 2.1 Caso especial "0" -> "000" (si el excel se comió los ceros)
                    elif b_look == "0" and "000" in cache_bodegas_cod:
                         bodega_obj = cache_bodegas_cod["000"]
                    # 3. Nombre
                    elif b_look in cache_bodegas:
                         bodega_obj = cache_bodegas[b_look]
                    else:
                         raise ValidationError(f"Bodega '{bodega_val}' no encontrada (por ID, Código o Nombre).")
                
                if not bodega_obj:
                    raise ValidationError("Falta especificar Bodega (en archivo o selección global).")

                # --- Resolver Tercero ---
                tercero_val = str(get_val("tercero", "")).strip()
                tercero_obj = default_tercero_obj
                
                if tercero_val:
                    t_look = tercero_val.lower()
                    if tercero_val in cache_terceros_id:
                        tercero_obj = cache_terceros_id[tercero_val]
                    elif t_look in cache_terceros_cod: # Codigo
                        tercero_obj = cache_terceros_cod[t_look]
                    elif t_look in cache_terceros: # Nombre
                         tercero_obj = cache_terceros[t_look]
                    else:
                         raise ValidationError(f"Tercero '{tercero_val}' no encontrado.")
                
                if not tercero_obj:
                    raise ValidationError("Falta especificar Tercero (en archivo o selección global).")

                # --- Otros campos ---
                prov_nombre = str(get_val("proveedor_nombre", "")).strip()
                proveedor_obj = None
                if prov_nombre:
                    p_upper = prov_nombre.upper()
                    if p_upper in cache_proveedores:
                        proveedor_obj = cache_proveedores[p_upper]
                    else:
                        # Crear proveedor on the fly si no existe
                        proveedor_obj = Proveedor.objects.create(nombre=prov_nombre)
                        cache_proveedores[p_upper] = proveedor_obj

                observacion = str(get_val("observacion", "")).strip()
                color = str(get_val("color", "")).strip()
                factura = str(get_val("factura", "")).strip()
                unidad_medida = str(get_val("unidad_medida", "")).strip().upper() # Capturar unidad

                # --- Lógica de Creación / Actualización ---
                insumo = Insumo.objects.filter(codigo=codigo).first()
                insumo_existed = True

                if not insumo:
                    insumo_existed = False
                    insumo = Insumo.objects.create(
                        codigo=codigo,
                        nombre=nombre or f"Insumo {codigo}",
                        proveedor=proveedor_obj,
                        bodega=bodega_obj,
                        color=color,
                        factura=factura,
                        observacion=observacion,
                        referencia=codigo,
                        unidad_medida=unidad_medida # Guardar unidad
                    )
                    registrar_movimiento_sin_afectar_stock(
                        insumo=insumo, tercero=tercero_obj, tipo="CREACION",
                        cantidad=Decimal("0"), costo_unitario=costo_unitario,
                        bodega=bodega_obj, observacion="Auto-creado Import"
                    )
                else:
                    # Si ya existe, actualizamos metadata básica pero NO stock
                    if nombre: insumo.nombre = nombre
                    if proveedor_obj: insumo.proveedor = proveedor_obj
                    if color: insumo.color = color
                    if unidad_medida: insumo.unidad_medida = unidad_medida # Actualizar unidad si viene
                    insumo.save()

                # Registrar entrada si viene cantidad > 0, SEA NUEVO O EXISTENTE
                # (Interpretando la columna como "Cantidad a sumar")
                if cantidad_entrada > 0:
                    # ✅ Si el insumo YA existía, usamos su costo actual para evitar distorsiones por error en Excel
                    # Salvo que el insumo tenga costo 0, entonces intentamos usar el del Excel.
                    costo_para_movimiento = costo_unitario
                    if insumo_existed and insumo.costo_unitario > 0:
                         costo_para_movimiento = insumo.costo_unitario

                    mov = aplicar_movimiento_insumo(
                        insumo=insumo,
                        tercero=tercero_obj,
                        tipo="ENTRADA",
                        cantidad=cantidad_entrada,
                        costo_unitario=costo_para_movimiento,
                        bodega=bodega_obj,
                        factura=factura,
                        observacion=observacion
                    )
                    movimientos_creados.append(mov.id)
                
                ok += 1

            except Exception as e:
                # Extraer mensaje limpio si es ValidationError de DRF
                msg = str(e)
                if hasattr(e, 'detail'):
                    d = e.detail
                    if isinstance(d, list):
                        # [ErrorDetail(string='Msg', code='invalid')]
                        msg = " ".join([str(x) for x in d])
                    elif isinstance(d, dict):
                        # {'field': ['Error']}
                        msg = " | ".join([f"{k}: {' '.join([str(x) for x in v]) if isinstance(v, list) else str(v)}" for k, v in d.items()])
                    else:
                        msg = str(d)
                
                errores.append({"fila": i, "error": msg})

        return Response(
            {
                "ok": True,
                "procesadas_ok": ok,
                "errores": errores,
                "movimientos_ids": movimientos_creados,
            },
            status=status.HTTP_200_OK
        )

    @action(detail=False, methods=["get"], url_path="plantilla-terminado", renderer_classes=[XLSXRenderer])
    def plantilla_terminado(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "ProductoTerminado"

        headers = [
            "fecha", "bodega_id", "tercero_id", "observacion",
            "producto_sku", "talla", "cantidad", "costo_unitario",
        ]
        ws.append(headers)
        ws.append([
            "2025-12-27", 1, 1, "Ingreso inicial por Excel",
            "CAM-001", "M", "12.000", "45000.00",
        ])

        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()

        resp = Response(content, content_type=XLSXRenderer.media_type)
        resp["Content-Disposition"] = 'attachment; filename="plantilla_producto_terminado.xlsx"'
        return resp

    @action(detail=False, methods=["post"], url_path="importar-terminado")
    def importar_terminado(self, request):
        file = request.FILES.get("file")
        if not file:
            raise ValidationError({"file": "Debe enviar un archivo .xlsx en multipart/form-data con key 'file'."})

        try:
            wb = load_workbook(filename=file, data_only=True)
        except Exception as e:
            raise ValidationError({"file": f"No se pudo leer el Excel: {str(e)}"})

        ws = wb["ProductoTerminado"] if "ProductoTerminado" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValidationError({"file": "El Excel está vacío."})

        header = [str(x).strip() if x is not None else "" for x in rows[0]]
        required = ["fecha", "bodega_id", "tercero_id", "producto_sku", "cantidad"]
        missing = [h for h in required if h not in header]
        if missing:
            raise ValidationError({"headers": f"Faltan columnas requeridas: {missing}"})

        idx = {h: header.index(h) for h in header if h}

        ok = 0
        errores = []
        movimientos = []

        # cache de notas: key -> nota_id (guardamos ID para evitar objetos “fantasma” si una fila falla)
        notas_cache = {}  # key=(fecha,bodega_id,tercero_id,obs) -> nota_id

        for i, r in enumerate(rows[1:], start=2):
            try:
                # ✅ transacción por fila (si falla, no rompe el resto)
                with transaction.atomic():

                    fecha = _parse_date(r[idx["fecha"]], "fecha") or timezone.now().date()

                    bodega = Bodega.objects.get(id=int(r[idx["bodega_id"]]))
                    tercero = Tercero.objects.get(id=int(r[idx["tercero_id"]]))

                    obs = str(r[idx["observacion"]] or "").strip() if "observacion" in idx else ""
                    producto_sku = str(r[idx["producto_sku"]]).strip()

                    talla_txt = str(r[idx["talla"]] or "").strip() if "talla" in idx else ""
                    cantidad = _parse_decimal(r[idx["cantidad"]], "cantidad")
                    if cantidad is None or cantidad <= 0:
                        raise ValidationError({"cantidad": "Debe ser > 0"})

                    costo_unitario = _parse_decimal(r[idx["costo_unitario"]], "costo_unitario") if "costo_unitario" in idx else None
                    costo_unitario = (costo_unitario.quantize(Decimal("0.01")) if costo_unitario else Decimal("0.00"))

                    producto = Producto.objects.get(codigo_sku=producto_sku)

                    talla_obj = None
                    if talla_txt:
                        talla_obj = Talla.objects.filter(nombre=talla_txt).first()
                        if not talla_obj:
                            raise ValidationError({"talla": f"La talla '{talla_txt}' no existe. Créala antes o deja vacío."})

                    key = (str(fecha), bodega.id, tercero.id, obs)

                    nota_id = notas_cache.get(key)
                    if nota_id:
                        nota = NotaEnsamble.objects.get(id=nota_id)
                    else:
                        nota = NotaEnsamble.objects.create(
                            bodega=bodega,
                            tercero=tercero,
                            fecha_elaboracion=fecha,
                            observaciones=(obs or "Ingreso por importación Excel (producto terminado)")
                        )
                        notas_cache[key] = nota.id

                    # ✅ en vez de CREATE siempre, hacemos UPSERT: si existe, sumamos
                    det, created = NotaEnsambleDetalle.objects.get_or_create(
                        nota=nota,
                        producto=producto,
                        talla=talla_obj,
                        bodega_actual=bodega,
                        defaults={"cantidad": cantidad},
                    )
                    if not created:
                        det.cantidad = (Decimal(str(det.cantidad or 0)) + Decimal(str(cantidad))).quantize(Decimal("0.001"))
                        det.cantidad_disponible = (Decimal(str(det.cantidad_disponible or 0)) + Decimal(str(cantidad))).quantize(Decimal("0.001"))
                        det.save(update_fields=["cantidad", "cantidad_disponible"])

                    # stock global
                    datos = DatosAdicionalesProducto.objects.filter(producto=producto).first()
                    if not datos:
                        datos = DatosAdicionalesProducto.objects.create(
                            producto=producto,
                            referencia="N/A",
                            unidad=producto.unidad_medida or "",
                            stock=Decimal("0.000"),
                            stock_minimo=Decimal("0"),
                            descripcion="",
                            marca="",
                            modelo="",
                            codigo_arancelario="",
                        )
                    datos.stock = (Decimal(str(datos.stock or 0)) + Decimal(str(cantidad))).quantize(Decimal("0.001"))
                    datos.save(update_fields=["stock"])

                    mov = ProductoTerminadoMovimiento.objects.create(
                        fecha=timezone.now(),
                        bodega=bodega,
                        tercero=tercero,
                        tipo=ProductoTerminadoMovimiento.Tipo.INGRESO_EXCEL,
                        producto=producto,
                        talla=talla_obj,
                        cantidad=cantidad,
                        costo_unitario=costo_unitario,
                        saldo_global_resultante=datos.stock,
                        nota_ensamble=nota,
                        observacion=f"{obs} (fila {i})".strip(),
                    )

                    movimientos.append(mov.id)
                    ok += 1

            except Exception as e:
                errores.append({"fila": i, "error": str(e)})

        return Response(
            {
                "ok": True,
                "procesadas_ok": ok,
                "errores": errores,
                "movimientos_ids": movimientos,
                "notas_creadas": sorted(set(notas_cache.values())),
            },
            status=status.HTTP_200_OK
        )

    @action(detail=False, methods=["get"], url_path="kardex-terminado")
    def kardex_terminado(self, request):
        qs = ProductoTerminadoMovimiento.objects.select_related("producto", "talla", "tercero", "bodega", "nota_ensamble").all()

        sku = (request.query_params.get("sku") or "").strip()
        if sku:
            qs = qs.filter(producto__codigo_sku=sku)

        bodega_id = request.query_params.get("bodega_id")
        if bodega_id and str(bodega_id).isdigit():
            qs = qs.filter(bodega_id=int(bodega_id))

        tercero_id = request.query_params.get("tercero_id")
        if tercero_id and str(tercero_id).isdigit():
            qs = qs.filter(tercero_id=int(tercero_id))

        qs = qs.order_by("-fecha", "-id")[:200]  # límite seguro

        return Response(ProductoTerminadoMovimientoSerializer(qs, many=True).data)